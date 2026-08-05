"""Noise campaign endpoints — ingest, readings, summary.

The logger export this parses is the format BSA's meters produce: a header
row then one row per interval with a running number, a date, a time and the
dB(A) value — exactly the four columns of the Dragon Ball raw file. Column
detection is by header name with sensible fallbacks, so a file whose columns
are ordered differently, or a plain CSV, still loads. Timestamps are stored
naive local (KSA), per the locked rule.

Validation on ingest mirrors the air side's spirit: a level outside the
physically plausible range for an outdoor A-weighted measurement is flagged
invalid at ingest rather than silently kept, and the summary excludes it.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, time as dtime, timedelta
from typing import List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status

from audit import audit
from auth import current_username
from db import db, to_mongo
from models import Campaign, NoiseReading
from noise_calc import (CONSTRUCTION_CORRECTIONS, NOISE_LIMITS, assess,
                        build_noise_summary)

log = logging.getLogger(__name__)
router = APIRouter(tags=["noise"])

# An outdoor A-weighted level below 20 dB(A) is quieter than a recording
# studio and above 140 is past the threshold of pain — either is a meter or
# export fault, not an environment.
MIN_DB, MAX_DB = 20.0, 140.0


def _campaign_or_404_sync(doc: Optional[dict]) -> Campaign:
    if not doc:
        raise HTTPException(status_code=404, detail="Campaign not found")
    return Campaign(**doc)


def _parse_timestamp(dval, tval, dayfirst: bool = False
                     ) -> Optional[datetime]:
    """Combine the date and time cells, tolerating the shapes Excel and CSV
    exports actually produce. ``dayfirst`` settles slashed dates — decided
    once per file by ``_slash_order``, never guessed per row."""
    d = None
    if isinstance(dval, datetime):
        d = dval.date()
    elif isinstance(dval, str) and dval.strip():
        fmts = (("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y")
                if dayfirst else
                ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y"))
        for fmt in fmts:
            try:
                d = datetime.strptime(dval.strip().split()[0], fmt).date()
                break
            except ValueError:
                continue
    t = None
    if isinstance(tval, dtime):
        t = tval.replace(microsecond=0)
    elif isinstance(tval, datetime):
        t = tval.time().replace(microsecond=0)
    elif isinstance(tval, str) and tval.strip():
        raw = tval.strip().split(".")[0]
        for fmt in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"):
            try:
                t = datetime.strptime(raw, fmt).time()
                break
            except ValueError:
                continue
    if d is None and isinstance(tval, datetime):
        return tval.replace(microsecond=0)
    if d is None or t is None:
        return None
    return datetime.combine(d, t)


def _find_columns(header: List) -> Tuple[Optional[int], Optional[int],
                                         Optional[int]]:
    """(date_col, time_col, db_col) by header text, with fallbacks."""
    date_i = time_i = db_i = None
    for i, h in enumerate(header):
        t = str(h or "").strip().lower()
        if not t:
            continue
        if db_i is None and ("db" in t or "laeq" in t or "leq" in t
                             or "spl" in t or "level" in t):
            db_i = i
        elif time_i is None and "time" in t:
            time_i = i
        elif date_i is None and "date" in t:
            date_i = i
    if db_i is None and len(header) >= 4:
        db_i = 3                       # the logger's fixed fourth column
    if date_i is None and len(header) >= 2:
        date_i = 1
    if time_i is None and len(header) >= 3:
        time_i = 2
    return date_i, time_i, db_i


def _slash_order(date_cells: List[str]) -> bool:
    """True when slashed dates in this file are day-first.

    A date like 7/10/2026 is genuinely ambiguous — July 10 to an American
    meter, 7 October to almost everyone else — and guessing wrong quietly
    moves a whole survey three months, which is exactly what happened to the
    first real upload. The order is decided for the file as a whole:

    * any first part over 12 proves day-first; any second part over 12
      proves month-first;
    * if every value is 12 or under, both readings are tried on the file's
      distinct dates and the one spanning the fewer days wins — a survey's
      dates are consecutive, and the wrong reading scatters them across
      months.
    """
    firsts, seconds, dates = set(), set(), set()
    for c in date_cells:
        parts = str(c).strip().split()[0].split("/")
        if len(parts) != 3:
            continue
        try:
            a, b = int(parts[0]), int(parts[1])
        except ValueError:
            continue
        firsts.add(a)
        seconds.add(b)
        dates.add((a, b, int(parts[2]) if parts[2].isdigit() else 0))
    if not dates:
        return False
    if any(a > 12 for a in firsts):
        return True
    if any(b > 12 for b in seconds):
        return False

    def span(dayfirst):
        vals = []
        for a, b, y in dates:
            try:
                vals.append(datetime(y, b, a) if dayfirst
                            else datetime(y, a, b))
            except ValueError:
                return None
        return (max(vals) - min(vals)).days
    s_df, s_mf = span(True), span(False)
    if s_df is None:
        return False
    if s_mf is None:
        return True
    return s_df < s_mf


def _parse_rows(raw: bytes, filename: str) -> List[Tuple[datetime, float]]:
    name = (filename or "").lower()
    rows: List[List] = []
    if name.endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True,
                                    data_only=True)
        ws = wb[wb.sheetnames[0]]
        for r in ws.iter_rows(values_only=True):
            rows.append(list(r))
    elif name.endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(file_contents=raw)
        sh = wb.sheet_by_index(0)
        for i in range(sh.nrows):
            row = []
            for j in range(sh.ncols):
                c = sh.cell(i, j)
                if c.ctype == xlrd.XL_CELL_DATE:
                    row.append(datetime(*xlrd.xldate_as_tuple(
                        c.value, wb.datemode)))
                else:
                    row.append(c.value)
            rows.append(row)
    else:                              # CSV / TSV
        import csv
        text = raw.decode("utf-8-sig", "replace")
        dialect = csv.excel_tab if "\t" in text.splitlines()[0] else csv.excel
        for r in csv.reader(io.StringIO(text), dialect):
            rows.append(list(r))
    if not rows:
        return []

    # The header is the first row containing any letters.
    h_idx = 0
    for i, r in enumerate(rows[:5]):
        if any(isinstance(c, str) and any(ch.isalpha() for ch in c)
               for c in r):
            h_idx = i
            break
    date_i, time_i, db_i = _find_columns(rows[h_idx])
    if db_i is None:
        raise HTTPException(
            status_code=422,
            detail=("Could not find a sound-level column. Expected columns "
                    "like: No. / Date / Time / dB."))

    dayfirst = _slash_order(
        [r[date_i] for r in rows[h_idx + 1:]
         if date_i is not None and date_i < len(r)
         and isinstance(r[date_i], str) and "/" in r[date_i]][:5000])

    out: List[Tuple[datetime, float]] = []
    for r in rows[h_idx + 1:]:
        if db_i >= len(r):
            continue
        try:
            level = float(str(r[db_i]).strip())
        except (TypeError, ValueError):
            continue
        ts = _parse_timestamp(r[date_i] if date_i is not None
                              and date_i < len(r) else None,
                              r[time_i] if time_i is not None
                              and time_i < len(r) else None,
                              dayfirst=dayfirst)
        if ts is None:
            continue
        out.append((ts, level))
    return out


def _fix_date_rollovers(pairs: List[Tuple[datetime, float]]
                        ) -> List[Tuple[datetime, float]]:
    """Logger exports often carry one date per block while the time runs past
    midnight — the Dragon Ball file does exactly this. Whenever the time goes
    backwards without the date advancing, roll the date forward."""
    if not pairs:
        return pairs
    fixed = [pairs[0]]
    offset = timedelta(0)
    for prev, cur in zip(pairs, pairs[1:]):
        ts = cur[0] + offset
        if ts < fixed[-1][0] - timedelta(minutes=5):
            offset += timedelta(days=1)
            ts = cur[0] + offset
        fixed.append((ts, cur[1]))
    return fixed


@router.post("/campaigns/{campaign_id}/noise-readings",
             status_code=status.HTTP_201_CREATED)
async def upload_noise_readings(campaign_id: str,
                                file: UploadFile = File(...),
                                user: str = Depends(current_username)):
    doc = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
    campaign = _campaign_or_404_sync(doc)
    if campaign.campaign_type != "noise":
        raise HTTPException(
            status_code=422,
            detail="This is not a noise campaign — upload analyser data "
                   "through the ordinary readings upload instead.")

    raw = await file.read()
    if len(raw) > 30 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File exceeds 30 MB")
    pairs = _fix_date_rollovers(_parse_rows(raw, file.filename or ""))
    if not pairs:
        raise HTTPException(
            status_code=422,
            detail="No readable rows found. Expected columns: "
                   "No. / Date / Time / dB.")

    # Replace, never merge: re-uploading is how a wrong file gets corrected.
    await db.noise_readings.delete_many({"campaign_id": campaign_id})

    docs, flagged = [], 0
    for ts, level in pairs:
        valid = MIN_DB <= level <= MAX_DB
        if not valid:
            flagged += 1
        docs.append(to_mongo(NoiseReading(
            campaign_id=campaign_id, timestamp=ts, laeq=level, valid=valid,
            invalidation_reason=None if valid else
            f"Level {level:.1f} dB(A) outside plausible range "
            f"({MIN_DB:.0f}–{MAX_DB:.0f})").model_dump()))
    if docs:
        await db.noise_readings.insert_many(docs)

    await db.campaigns.update_one(
        {"id": campaign_id},
        {"$set": {"status": "ingested",
                  "updated_at": datetime.utcnow()}})
    await audit("noise.upload", "campaign", campaign_id, user,
                {"rows": len(docs), "auto_flagged": flagged,
                 "filename": file.filename})
    return {"rows": len(docs), "auto_flagged": flagged,
            "first": docs[0]["timestamp"] if docs else None,
            "last": docs[-1]["timestamp"] if docs else None}


@router.get("/campaigns/{campaign_id}/noise-readings")
async def list_noise_readings(campaign_id: str, limit: int = 2000,
                              skip: int = 0):
    docs = await db.noise_readings.find(
        {"campaign_id": campaign_id}, {"_id": 0}) \
        .sort("timestamp", 1).skip(skip).to_list(length=min(limit, 5000))
    total = await db.noise_readings.count_documents(
        {"campaign_id": campaign_id})
    return {"total": total, "items": docs}


@router.delete("/campaigns/{campaign_id}/noise-readings", status_code=204)
async def clear_noise_readings(campaign_id: str,
                               user: str = Depends(current_username)):
    await db.noise_readings.delete_many({"campaign_id": campaign_id})
    await audit("noise.clear", "campaign", campaign_id, user, {})


@router.get("/campaigns/{campaign_id}/noise-summary")
async def noise_summary(campaign_id: str):
    doc = await db.campaigns.find_one({"id": campaign_id}, {"_id": 0})
    campaign = _campaign_or_404_sync(doc)
    readings = await db.noise_readings.find(
        {"campaign_id": campaign_id}, {"_id": 0}).sort("timestamp", 1) \
        .to_list(length=None)
    if not readings:
        raise HTTPException(status_code=400,
                            detail="No noise readings uploaded")
    s = build_noise_summary(readings, campaign.monitoring_start,
                            campaign.monitoring_end,
                            campaign.day_start_hour, campaign.day_end_hour)
    verdicts = assess(s, campaign.noise_category)
    limit = NOISE_LIMITS.get(campaign.noise_category)
    return {
        "category": campaign.noise_category,
        "category_label": limit.label if limit else campaign.noise_category,
        "limits": ({"day": limit.day_db, "night": limit.night_db}
                   if limit and limit.day_db is not None else None),
        "corrections": [{"period": p, "db": c}
                        for p, c in CONSTRUCTION_CORRECTIONS],
        "stats": {
            "laeq_t": s.laeq_t, "l_day": s.l_day, "l_night": s.l_night,
            "la10": s.la10, "la50": s.la50, "la90": s.la90,
            "lmax": s.lmax, "lmin": s.lmin,
            "lmax_at": s.lmax_at,
        },
        "capture": {"total": s.total_records, "valid": s.valid_records,
                    "invalid": s.invalid_records,
                    "expected": s.expected_records,
                    "pct": s.data_capture_pct},
        "day_window": {"start": s.day_start_hour, "end": s.day_end_hour,
                       "day_records": s.day_records,
                       "night_records": s.night_records},
        "verdicts": [{"period": v.period, "measured": v.measured,
                      "limit": v.limit, "margin": v.margin,
                      "status": v.status, "text": v.text}
                     for v in verdicts],
    }
