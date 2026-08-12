"""Read BSA's own Certificate of Analysis workbook.

The laboratory already produces a file for every job — the F-BR-29 Final
Test Report, or the non-accredited variant of it. Asking them to produce a
second file in a different shape for this system is work the system
created, not work the job required. This reads the file they already make.

The layout is not fixed. Column positions move between the blank template
and the filled ones, the header block sits at different rows, and the
non-accredited version has no MU or LOQ columns at all. So nothing is read
from a fixed cell reference: the header row is found by looking for the
word "parameter", the columns are mapped by their headings, and the sample
metadata is found by matching labels.

One thing is deliberately not read: **the limit columns.** The sheet
carries its own NCEC limits, and on the SAJCO certificate the F1 limit is
printed as 120 mg/kg where Appendix (1) gives 210. Importing the sheet's
limits would import that error and hide it behind a system that looks
authoritative. Limits come from the regulation, every time.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

log = logging.getLogger(__name__)

# Sheets that never hold results.
SKIP_SHEETS = {"cover", "summary", "notes", "index", "contents"}

# Column headings, in the spellings that appear across the templates.
COL_PATTERNS: List[Tuple[str, Tuple[str, ...]]] = [
    ("parameter", ("parameter(s)", "parameters", "parameter")),
    ("result", ("result(s)", "results", "result", "value")),
    ("unit", ("unit(s)", "units", "unit")),
    ("method", ("method(s)", "methods", "method", "test method")),
    ("mu", ("mu%", "mu %", "mu", "uncertainty", "measurement uncertainty")),
    ("loq", ("loq", "lod", "limit of quantification")),
]

# Headings that mark the limit columns. Found so they can be skipped, and
# reported so a reviewer knows the sheet carried limits that were ignored.
LIMIT_PATTERNS = ("limit", "limits", "ncec", "standard", "guideline",
                  "coarse soil", "soft soil")

# Metadata labels in the header block, mapped to the field they fill.
META_PATTERNS: List[Tuple[str, Tuple[str, ...]]] = [
    ("sample_code", ("sampling id", "sample code", "sample id",
                     "sample  code")),
    ("client", ("client name", "customer name", "client")),
    ("report_id", ("report id", "report no", "report number")),
    ("sampled_at", ("sampling date", "date of sampling")),
    ("received_at", ("receiving date", "date received", "received date")),
    ("reported_at", ("reporting date", "report date", "date of report")),
    ("site", ("name station", "sampling site", "sampling location",
              "station name", "site")),
    ("description", ("sample  description", "sample description",
                     "description")),
    ("coc", ("coc no.", "coc no", "coc number", "chain of custody")),
    ("sampled_by", ("sampling by", "sampled by", "sampling method")),
]

# Words in the sample description that identify the soil context. Read as a
# suggestion only: the parsed value is offered to the operator to confirm,
# never applied silently, because it decides which column of Appendix (1)
# every result is judged against.
PARTICLE_WORDS = {"coarse": "coarse", "sand": "coarse", "gravel": "coarse",
                  "soft": "soft", "silt": "soft", "clay": "soft",
                  "fine": "soft"}
DEPTH_WORDS = {"surface": "topsoil", "topsoil": "topsoil", "top soil":
               "topsoil", "subsurface": "subsurface", "sub-surface":
               "subsurface", "deep": "subsurface"}


@dataclass
class ParsedResult:
    parameter: str
    raw_value: Optional[str] = None
    value: Optional[float] = None
    below_loq: bool = False
    unit: Optional[str] = None
    method: Optional[str] = None
    mu_percent: Optional[float] = None
    loq: Optional[float] = None


@dataclass
class ParsedSample:
    sheet: str
    sample_code: Optional[str] = None
    client: Optional[str] = None
    report_id: Optional[str] = None
    site: Optional[str] = None
    description: Optional[str] = None
    sampled_by: Optional[str] = None
    coc: Optional[str] = None
    sampled_at: Optional[datetime] = None
    received_at: Optional[datetime] = None
    reported_at: Optional[datetime] = None
    # Suggestions read from the sample description, for the operator to
    # confirm. Never applied without confirmation.
    suggested_particle_size: Optional[str] = None
    suggested_depth: Optional[str] = None
    results: List[ParsedResult] = field(default_factory=list)
    notes: List[str] = field(default_factory=list)
    # Parameter names found in the table, whether or not they carried a
    # result. Used to tell a blank template apart from an unreadable sheet.
    parameter_names_seen: int = 0


@dataclass
class ParsedWorkbook:
    samples: List[ParsedSample] = field(default_factory=list)
    skipped_sheets: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    # True where the sheet carried its own limit columns, which were read
    # and discarded.
    limits_ignored: bool = False


def _txt(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return " ".join(str(v).split()).strip()


def _norm(v: Any) -> str:
    return _txt(v).lower().rstrip(":").strip()


def _match(text: str, patterns: Tuple[str, ...]) -> bool:
    t = _norm(text)
    return any(t == p or t.startswith(p) for p in patterns)


def _to_number(v: Any) -> Tuple[Optional[float], Optional[str], bool]:
    """Read a result cell: (value, as-written, below the quantification limit).

    A cell written "<0.001" keeps its text and is flagged. It is never
    turned into 0.001 and never into zero. Anything unparseable keeps its
    text with no value, so it prints as the laboratory wrote it and is not
    judged.
    """
    if v is None:
        return None, None, False
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v), None, False
    raw = _txt(v)
    if not raw or _norm(raw) in {"-", "na", "n/a", "nd", "\u2014", "\u00ad"}:
        return None, (raw or None), False
    cleaned = raw.replace(",", "").replace("\u2212", "-").replace("\u00b1", "")
    below = cleaned.startswith("<")
    if below:
        cleaned = cleaned[1:].strip()
    cleaned = cleaned.replace("%", "").strip()
    try:
        return float(cleaned), raw, below
    except ValueError:
        return None, raw, False


def _to_dt(v: Any) -> Optional[datetime]:
    """Dates are not guessed.

    The certificates carry both "29/7/2026" and "2026-03-08", and read the
    wrong way round the second is 8 March rather than 3 August. Only
    unambiguous forms are accepted; anything else is left unset and the
    operator fills it in, because a plausible wrong date on a legal record
    is worse than a blank one.
    """
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    raw = _txt(v)
    if not raw:
        return None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", raw)
    if m:
        y, a, b = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(y, a, b)
        except ValueError:
            return None
    m = re.match(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})$", raw)
    if m:
        first, second, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        # Only accept it where one of the two can only be the day.
        if first > 12 and second <= 12:
            try:
                return datetime(year, second, first)
            except ValueError:
                return None
        if second > 12 and first <= 12:
            try:
                return datetime(year, first, second)
            except ValueError:
                return None
        # Both under 13 — genuinely ambiguous. Left unset.
        return None
    return None


def _row_values(ws, row: int) -> List[Any]:
    return [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]


def _find_header(ws) -> Optional[Tuple[int, Dict[str, int], bool]]:
    """Locate the results table header and map its columns.

    Returns (row index, {field: column index}, whether limit columns exist).
    """
    for row in range(1, min(ws.max_row, 60) + 1):
        values = _row_values(ws, row)
        cols: Dict[str, int] = {}
        has_limits = False
        for idx, v in enumerate(values):
            label = _norm(v)
            if not label:
                continue
            if any(p in label for p in LIMIT_PATTERNS):
                has_limits = True
                continue
            for key, patterns in COL_PATTERNS:
                if key in cols:
                    continue
                if _match(label, patterns):
                    cols[key] = idx
                    break
        if "parameter" in cols and "result" in cols:
            return row, cols, has_limits
    return None


def _read_meta(ws, header_row: int, sample: ParsedSample) -> None:
    """Read the header block above the results table.

    Labels and values are not in fixed columns, so each label is matched and
    the value taken from the next non-empty cell to its right.
    """
    for row in range(1, header_row):
        values = _row_values(ws, row)
        for idx, cell in enumerate(values):
            label = _norm(cell)
            if not label:
                continue
            for key, patterns in META_PATTERNS:
                if getattr(sample, key, None):
                    continue
                if not _match(label, patterns):
                    continue
                for j in range(idx + 1, len(values)):
                    raw = values[j]
                    if raw is None or not _txt(raw):
                        continue
                    if _norm(raw) in {p for _, ps in META_PATTERNS
                                      for p in ps}:
                        break
                    if key in ("sampled_at", "received_at", "reported_at"):
                        parsed = _to_dt(raw)
                        if parsed:
                            setattr(sample, key, parsed)
                        else:
                            sample.notes.append(
                                f"{key.replace('_', ' ')} '{_txt(raw)}' could "
                                f"not be read unambiguously and was left "
                                f"blank")
                    else:
                        setattr(sample, key, _txt(raw))
                    break
                break


def _suggest_context(sample: ParsedSample) -> None:
    text = " ".join(x for x in (sample.description, sample.sampled_by) if x)
    low = text.lower()
    for word, value in PARTICLE_WORDS.items():
        if word in low:
            sample.suggested_particle_size = value
            break
    for word, value in DEPTH_WORDS.items():
        if word in low:
            sample.suggested_depth = value
            break


def _read_results(ws, header_row: int, cols: Dict[str, int],
                  sample: ParsedSample) -> None:
    """Read the rows below the header until the table runs out.

    A row with a parameter name but no result is a group heading — "TOTAL
    PETROLEUM HYDROCARBON (TPH)", "Microbiology parametrs". Those are
    skipped rather than imported as a parameter with no value, which would
    put a phantom row in the report.
    """
    blanks = 0
    names_seen = 0
    for row in range(header_row + 1, ws.max_row + 1):
        values = _row_values(ws, row)
        name = _txt(values[cols["parameter"]]
                    if cols["parameter"] < len(values) else None)
        result_raw = (values[cols["result"]]
                      if cols["result"] < len(values) else None)
        if not name and result_raw is None:
            blanks += 1
            # The signature block sits several blank rows below the table.
            if blanks >= 8:
                break
            continue
        blanks = 0
        if not name:
            continue
        low = _norm(name)
        if low.startswith(("review", "approved", "date", "sign", "note",
                           "opinion", "statements")):
            break
        names_seen += 1
        value, raw, below = _to_number(result_raw)
        if value is None and raw is None:
            # A heading row inside the table.
            continue

        def _col(key: str):
            idx = cols.get(key)
            if idx is None or idx >= len(values):
                return None
            return values[idx]

        mu_val, _, _ = _to_number(_col("mu"))
        loq_val, _, _ = _to_number(_col("loq"))
        sample.results.append(ParsedResult(
            parameter=name,
            raw_value=raw,
            value=value,
            below_loq=below,
            unit=_txt(_col("unit")) or None,
            method=_txt(_col("method")) or None,
            mu_percent=mu_val,
            loq=loq_val,
        ))
    sample.parameter_names_seen = names_seen


def parse_coa_workbook(data: bytes) -> ParsedWorkbook:
    """Parse a Certificate of Analysis workbook into samples and results.

    One sheet is one sample, which is how the laboratory already organises
    the file. A sheet with no recognisable results table is skipped and
    named, rather than silently ignored.
    """
    out = ParsedWorkbook()
    from io import BytesIO
    wb = load_workbook(BytesIO(data), data_only=True, read_only=False)
    for name in wb.sheetnames:
        if _norm(name) in SKIP_SHEETS:
            out.skipped_sheets.append(name)
            continue
        ws = wb[name]
        found = _find_header(ws)
        if not found:
            out.skipped_sheets.append(name)
            out.warnings.append(
                f"Sheet '{name}': no results table found. A header row "
                f"containing 'Parameter' and 'Result' is what identifies it.")
            continue
        header_row, cols, has_limits = found
        if has_limits:
            out.limits_ignored = True
        sample = ParsedSample(sheet=name)
        _read_meta(ws, header_row, sample)
        _suggest_context(sample)
        _read_results(ws, header_row, cols, sample)
        if not sample.results:
            out.skipped_sheets.append(name)
            if sample.parameter_names_seen:
                out.warnings.append(
                    f"Sheet '{name}': {sample.parameter_names_seen} parameter "
                    f"name(s) found but every result is empty. This looks "
                    f"like a blank template rather than a completed "
                    f"certificate.")
            else:
                out.warnings.append(
                    f"Sheet '{name}': a results table was found but no rows "
                    f"could be read from it.")
            continue
        if not sample.sample_code:
            sample.sample_code = f"Sheet {name}"
            sample.notes.append(
                "No sample code found on the sheet; the sheet name was used. "
                "Rename it before issuing.")
        for label, value in (("Receiving", sample.received_at),
                             ("Reporting", sample.reported_at)):
            if value and sample.sampled_at and value < sample.sampled_at:
                sample.notes.append(
                    f"{label} date ({value:%d %B %Y}) is before the sampling "
                    f"date ({sample.sampled_at:%d %B %Y}). One of the two is "
                    f"written the wrong way round — check before issuing.")
        if "mu" not in cols:
            sample.notes.append(
                "No measurement uncertainty column on this sheet. A guard "
                "band cannot be applied to these results.")
        out.samples.append(sample)
    return out
